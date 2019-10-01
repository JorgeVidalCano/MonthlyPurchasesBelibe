## Script to get the monthly purchases by country
# -*- coding: utf-8 -*-

import psycopg2
from datetime import datetime
from config import config
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def commonRows(row, salto_2, factura_rectificativa, factura, fecha, cliente):
	'''Common rows in the excel_creation function'''
	newRow = row+salto_2
	sheet['A{}'.format(newRow)] = factura_rectificativa
	sheet['B{}'.format(newRow)] = factura
	sheet['C{}'.format(newRow)] = fecha
	sheet['D{}'.format(newRow)] = cliente.split(",")[0]

def print10row(row, salto_2, precio_base):
	'''Common rows in the excel_creation function'''
	newRow = row+salto_2
	sheet['E{}'.format(newRow)] = precio_base
	sheet['F{}'.format(newRow)] = round(precio_base*1.1 - precio_base, 2) #%10
	sheet['G{}'.format(newRow)] = 0 #base21%
	sheet['H{}'.format(newRow)] = 0 #21%

def print21row(row, salto_2, precio_base):
	'''Common rows in the excel_creation function'''
	newRow = row+salto_2
	sheet['E{}'.format(newRow)] = 0 #base10%
	sheet['F{}'.format(newRow)] = 0 #10%
	sheet['G{}'.format(newRow)] = precio_base
	sheet['H{}'.format(newRow)] = round(precio_base*1.21 - precio_base, 2) #%21

def excel_creation(data, country):
	'''Creates the excel'''
	global sheet
	sheet = wb.active
	sheet = wb.create_sheet(country, 0)

	# Column's name
	sheet['A1'] = 'Factura rectificativa'
	sheet.column_dimensions['A'].width = 15
	sheet['B1'] = 'Factura origen'
	sheet.column_dimensions['B'].width = 13
	sheet['C1'] = 'Fecha Devolucion'
	sheet.column_dimensions['C'].width = 14
	sheet['D1'] = 'Cliente'
	sheet.column_dimensions['D'].width = 30
	sheet['E1'] = 'Base10'
	sheet['F1'] = 'Iva 10%'
	sheet['G1'] = 'Base21'
	sheet['H1'] = 'Iva 21%'
	sheet['I1'] = 'Ajuste'
	sheet['J1'] = 'Total'


	'''
	Data is the info retrieved by the database
	Sheet[] prints that data to the excel
	
	'''
	for row in range(0, len(data)+1):

		# data[row][0] and so on, selects an specific piece of data from the ddbb. (The data is given as a tuple)
		fecha = data[row][0]
		factura = data[row][1]
		factura_rectificativa = data[row][2]
		
		cliente = data[row][3]
		iva = data[row][4]
		precio_base = round(float(data[row][5]), 2)
		salto_2 = 2

		if row == 0:
			commonRows(row, salto_2, factura_rectificativa, factura, fecha,cliente) ## prints the string data
			if str(iva) == '56': #%10
				print10row(row, salto_2, precio_base)
			elif str(iva) == '1': #%21
				print21row(row, salto_2, precio_base)
		else:
			if factura == data[row-1][1]:
				if row == 1:
					if factura == data[row-1][1]: # Compares the factura with the one before. I had to hard coded the first and second line (of the excel) otherwise I got sheet[A0] or sheet[A-1]
						sheet['E{}'.format(row+1)] = precio_base # Base10
						sheet['F{}'.format(row+1)] = round(precio_base*1.1 - precio_base, 2)
					else:
						if str(iva) == 'None':
							sheet['I{}'.format(row+1)] = precio_base
						else:
							sheet['E{}'.format(row+1)] = precio_base, 2 # Base10
							sheet['F{}'.format(row+1)] = round(precio_base*1.1 - precio_base, 2)
				else:
					if factura == data[row-2][1]: # Compares the factura with the one before. I had to hard coded the first and second line (of the excel) otherwise I got sheet[A0] or sheet[A-1]
						sheet['I{}'.format(row)] = precio_base # Base10
					else:
						if str(iva) == 'None':
							sheet['I{}'.format(row+1)] = precio_base
						else:
							sheet['E{}'.format(row+1)] = precio_base # Base10
							sheet['F{}'.format(row+1)] = round(precio_base*1.1 - precio_base, 2)
			else:
				commonRows(row, salto_2, factura_rectificativa, factura, fecha,cliente) ## prints the string data
				if str(iva) == '56': #%10
					print10row(row, salto_2, precio_base) # Prints the int data
				elif str(iva) == '1': #%21
					print21row(row, salto_2, precio_base) # Prints the int data
			
def deleteEmptyRowsAndSumRows(country, fileName):

	pais = {69:"Espana", 110:"Italia", 21:"Belgica", 76:"Francia"}
	
	wb = load_workbook('{}.xlsx'.format(fileName))
	sheet = wb[pais[country]]
	ajustRow = 0
	first = True
	firstReturn = 0
	fillcolor = PatternFill(start_color="E2F1C1", end_color="E2F1C1", fill_type = "solid")

	for readRow in range(1, sheet.max_row+1):
		'''readRow reads the lines normally, if there is a gap, then it substract 1 to writeRow otherwise it won't remove white spaces.'''
		writeRow = readRow + ajustRow

		if str(sheet.cell(row=readRow, column=1).value).startswith('R') and first: ## Selects the start line of the returned orders
			firstReturn = writeRow
			first = False

		if sheet.cell(row=readRow, column=1).value is not None:
			
			for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']: ## Rewrites all lines yet again but with no white space between lines
				
				try:
					if str(sheet.cell(row=readRow, column=1).value).startswith('R'):
						fillcolor = PatternFill(start_color="EED5D2", end_color="EED5D2", fill_type = "solid")
				except:
					pass

				sheet['{}{}'.format(letter, writeRow)] = sheet['{}{}'.format(letter, readRow)].value ## just in case			
				sheet['{}{}'.format(letter, writeRow)].fill = fillcolor

			try: ## End row sum. Needs the try because a NoneType sums happen all the time
				if readRow == 1:
					sheet['J1'] = 'Total'
					sheet['J1'].fill = fillcolor
				else:
					sheet['J{}'.format(writeRow)] = sheet['E{}'.format(readRow)].value + sheet['F{}'.format(readRow)].value + sheet['G{}'.format(readRow)].value + sheet['H{}'.format(readRow)].value # Total sum
					sheet['J{}'.format(writeRow)].fill = fillcolor
			except:
				pass
			
		else: # if the line is empty, it substracts one the writeRow
			ajustRow -= 1
		
	for clearRow in range(writeRow+1, readRow+1): ## Cleans the remaining cells because it uses the same excel. It's just the way it is
		for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
			sheet['{}{}'.format(letter, clearRow)] = ' '

	if firstReturn == 0: ## In case there are no returns
		firstReturn = readRow

	for letter in ['D', 'E', 'F', 'G', 'H', 'J']:

		if letter == 'D':
			sheet['{}{}'.format(letter, writeRow+3)] = 'Ventas'.format(letter, letter, firstReturn-1) ## tag "Ventas"
			sheet['{}{}'.format(letter, writeRow+4)] = 'Abonos'.format(letter, letter, firstReturn-1) ## tag "Abonos"	
		else:
			sheet['{}{}'.format(letter, writeRow+3)] = '= SUM({}2:{}{})'.format(letter, letter, firstReturn-1) ## sum invoices
			sheet['{}{}'.format(letter, writeRow+4)] = '= -SUM({}{}:{}{})'.format(letter, firstReturn, letter, writeRow) # sum returns


	wb.save("{}.xlsx".format(fileName))

def connect(dateFrom, dateTo, country):
	"""Connexion to the database"""
	pais = {69:"Espana", 110:"Italia", 21:"Belgica", 76:"Francia"}
	conn = None
	dateFrom = datetime.strptime(dateFrom, '%Y-%m-%d')
	dateTo = datetime.strptime(dateTo, '%Y-%m-%d')

	try:
		params = config()
		print("Connecting to the database")
		conn = psycopg2.connect(**params)
		cur = conn.cursor()
		print("Connection suscessful")

		cur.execute("""
			SELECT
				to_char(ai.date_invoice, 'DD-MM-YYYY') AS fecha_devolucion,
				ai.origin AS factura_origen,
				ai.invoice_number AS factura_rectificativa,
				rp.display_name As nombre_cliente,
				ait.tax_id,
				sum(ail.price_subtotal) AS precio_base,
				ai.amount_tax
			FROM
				account_invoice AS ai
			LEFT JOIN res_partner AS rp ON (ai.partner_id = rp.id)
			LEFT JOIN account_invoice_line AS ail ON (ail.invoice_id = ai.id)
			LEFT JOIN account_invoice_line_tax AS ait ON (ait.invoice_line_id = ail.id)
			WHERE
			ai.date_invoice BETWEEN '{}' AND '{}' 
			AND rp.country_id = '{}'
			GROUP BY
				ai.ID,
				ait.tax_id,
				rp.display_name,
				rp.country_id
			ORDER BY
				ai.invoice_number,
  			    ait.tax_id	
				""".format(dateFrom, dateTo, country))
		result = cur.fetchall()
		
		cur.close()
		excel_creation(result, pais[country])

	except (Exception, psycopg2.DatabaseError) as error:
		print(error)
	finally:
		if conn is not None:
			conn.close()
			print("Database connection closed")


if __name__ == '__main__':
	month, year = datetime.now().strftime("%m-%Y").split('-')
	month = str(int(month) - 1)
	fileName = 'Ventas {}-{}'.format(month, year)

	lastMonthDay = {
		'1':31,
		'2':28,
		'3':31,
		'4':30,
		'5':31,
		'6':30,
		'7':31,
		'8':31,
		'9':30,
		'10':31,
		'11':30,
		'12':31
	}

	fromDate = ('{}-{}-1').format(year, month)
	toDate = ('{}-{}-{}').format(year, month, lastMonthDay[month])
	
	wb = Workbook()
	paises = [69, 110, 21, 76]
	
	for pais in (paises):
		connect(fromDate, toDate, pais)
	wb.save("{}.xlsx".format(fileName))

	for pais in (paises):
		deleteEmptyRowsAndSumRows(pais, fileName)
